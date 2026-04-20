# -*- coding: utf-8 -*-
import sys
print("Python version:", sys.version)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    print("openpyxl is available")
except ImportError as e:
    print("openpyxl not available:", e)
    sys.exit(1)

headers = ["题干", "题型", "难度", "选项A", "选项B", "选项C", "选项D", "正确答案", "知识点ID"]

questions = [
    ["中国第一艘载人航天飞船是？", "SINGLE", 1, "神舟五号", "神舟六号", "神舟七号", "神舟八号", "A", "1"],
    ["中国航天员第一人是谁？", "SINGLE", 1, "杨利伟", "聂海胜", "翟志刚", "刘伯明", "A", "1"],
    ["中国空间站的名称是？", "SINGLE", 1, "天宫", "天和", "问天", "梦天", "A", "1"],
    ["嫦娥工程主要探测的目标是？", "SINGLE", 1, "月球", "火星", "金星", "木星", "A", "2"],
    ["天问一号执行的任务是？", "SINGLE", 1, "火星探测", "月球探测", "空间站", "载人航天", "A", "2"],
    ["中国卫星导航系统的名称是？", "SINGLE", 1, "北斗", "GPS", "伽利略", "格洛纳斯", "A", "3"],
    ["长征系列火箭主要用于？", "SINGLE", 1, "运载卫星", "军事打击", "民航运输", "科学实验", "A", "4"],
    ["中国第一个月球探测器是？", "SINGLE", 1, "嫦娥一号", "嫦娥二号", "嫦娥三号", "嫦娥四号", "A", "2"],
    ["下列属于中国航天成就的有？", "MULTIPLE", 2, "神舟飞船", "天宫空间站", "北斗导航", "哈勃望远镜", "ABC", "1,3"],
    ["中国有哪些卫星导航系统？", "MULTIPLE", 2, "北斗", "GPS", "伽利略", "格洛纳斯", "ABCD", "3"],
    ["中国探月工程分为哪几个阶段？", "MULTIPLE", 2, "绕", "落", "回", "登", "ABC", "2"],
    ["长征火箭家族包括哪些型号？", "MULTIPLE", 2, "长征二号", "长征三号", "长征五号", "长征七号", "ABCD", "4"],
    ["中国空间站主要由哪些模块组成？", "MULTIPLE", 2, "天和核心舱", "问天实验舱", "梦天实验舱", "神舟飞船", "ABC", "1"],
    ["判断：中国是世界上第三个独立实现载人航天的国家。", "JUDGE", 1, "正确", "错误", "", "", "A", "1"],
    ["判断：北斗系统是中国自主研发的卫星导航系统。", "JUDGE", 1, "正确", "错误", "", "", "A", "3"],
    ["判断：嫦娥五号实现了月球采样返回。", "JUDGE", 1, "正确", "错误", "", "", "A", "2"],
    ["判断：天宫空间站是中国自主建设的空间站。", "JUDGE", 1, "正确", "错误", "", "", "A", "1"],
    ["判断：长征五号是中国推力最大的运载火箭。", "JUDGE", 1, "正确", "错误", "", "", "A", "4"],
    ["中国航天日是哪一天？", "FILL", 2, "", "", "", "", "4月24日", "1"],
    ["中国第一位进入太空的女航天员是？", "FILL", 2, "", "", "", "", "刘洋", "1"],
    ["天和核心舱于哪一年发射成功？", "FILL", 2, "", "", "", "", "2021年", "1"],
    ["请简述中国载人航天工程三步走战略。", "SUBJECTIVE", 3, "", "", "", "", "第一步：发射载人飞船；第二步：突破空间站关键技术；第三步：建造空间站", "1"],
    ["请简述北斗导航系统的主要功能。", "SUBJECTIVE", 3, "", "", "", "", "定位、导航、授时、短报文通信", "3"],
    ["请简述探月工程嫦娥五号的任务过程。", "SUBJECTIVE", 3, "", "", "", "", "发射、月球采样、上升器返回、对接、返回舱地球", "2"]
]

wb = Workbook()
ws = wb.active
ws.title = "航天知识题库"

# Header style
header_font = Font(bold=True)
header_fill = PatternFill(start_color="B4D7FF", end_color="B4D7FF", fill_type="solid")

# Write headers
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill

# Write questions
for row_idx, question in enumerate(questions, 2):
    for col_idx, value in enumerate(question, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# Auto adjust column width
for col in ws.columns:
    max_length = 0
    column = col[0].column
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    ws.column_dimensions[column].width = adjusted_width

output_path = r"d:\1\space-knowledge-platform\航天知识题库导入模板.xlsx"
wb.save(output_path)
print(f"Excel file generated successfully!")
print(f"File path: {output_path}")
print(f"Number of questions: {len(questions)}")
